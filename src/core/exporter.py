"""
品質モデルExcel出力モジュール

このモジュールは、AIによって再構成されたMarkdownテーブルを解析し、
正式な運用に適したプロフェッショナルなExcelフォーマット (.xlsx) で出力します。
メイリオフォント、列幅調整、ヘッダー固定などの書式設定を自動化します。

作成者: HAL (AI Agent)
作成日: 2026-04-10
"""

import os
import openpyxl
from openpyxl.styles import Alignment, Font

class Exporter:
    """
    最終的なMarkdownテーブルをパースし、Excelファイルとして出力するクラス。
    
    Attributes:
        output_dir (str): Excelファイルの出力先ディレクトリパス。
    """

    def __init__(self, output_dir: str):
        """
        Exporterクラスのコンストラクタ。
        出力ディレクトリの存在確認と自動作成を行います。
        
        Args:
            output_dir (str): Excelファイルの出力先パス。
        """
        self.output_dir = output_dir
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

    def export_to_excel(self, parent_char: str, subchar: str, markdown_table: str, author_prefix="(山本)"):
        """
        Markdown形式のテーブルを解析し、書式設定済みのExcelファイルに出力します。
        
        ファイル名は以下の形式で生成されます：
        (担当者)品質特性_品質副特性_再構成.xlsx

        Args:
            parent_char (str): 親となる品質特性名。
            subchar (str): 品質副特性名。
            markdown_table (str): AIから出力されたMarkdown形式のテーブルテキスト。
            author_prefix (str): ファイル名に付与する作成者の識別子。

        Returns:
            str: 生成されたExcelファイルの絶対パス。
        """
        parent_prefix = f"{parent_char}_" if parent_char else ""
        
        # 1. 安全なファイル名の生成
        # ファイル名に使えない文字をアンダースコアに置換
        safe_subchar = subchar.replace("/", "_").replace("\\", "_")
        safe_parent = parent_prefix.replace("/", "_")
        
        filename = f"{author_prefix}{safe_parent}{safe_subchar}_再構成.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        # 2. ワークブックの初期化
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = safe_subchar[:31]  # Excelシート名の31文字制限に対応

        # 3. リッチな固定ヘッダーの書き込み
        headers = ["測定項目", "測定基準", "記述例（適合例／非適合例）"]
        for col_idx, header_text in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header_text)

        # 4. Markdownテーブルのパースとデータ書き込み
        lines = markdown_table.strip().split("\n")
        row_idx = 2 # データ書き込み開始行 (2行目)
        table_started = False

        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # パイプ (|) で囲まれた行のみをテーブルの行として処理
            if line.startswith("|") and line.endswith("|"):
                # セパレーター行 (|---|:---|) を検知してスキップ
                content_only = line[1:-1].replace(" ", "").replace("|", "")
                if all(c in "-:" for c in content_only) and len(content_only) > 0:
                    continue
                
                # AIが回答に含めたヘッダー行をスキップ（自前で1行目を書いているため）
                if "測定項目" in line and "測定基準" in line:
                    table_started = True 
                    continue
                
                # テーブル開始前のメタデータ等はスキップ
                if not table_started:
                    pass

                # 列データの分解
                cols = [c.strip() for c in line[1:-1].split("|")]
                if len(cols) < 3:
                    continue
                
                # 最初の3列（測定項目, 測定基準, 記述例）を抽出
                target_cols = cols[:3]
                
                for col_idx, col_val in enumerate(target_cols, start=1):
                    # AIが付加した改行用プレースホルダを実際の改行コードに復元
                    cell_value = col_val.replace("【改行】", "\n")
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)
                row_idx += 1

        # 5. プロフェッショナルな書式設定 (DADA品質基準)
        # フォント設定: 日本語が読みやすい「メイリオ」を採用
        font_header = Font(name="メイリオ", size=10, bold=True)
        font_data = Font(name="メイリオ", size=10, bold=False)
        
        # 配置設定: ヘッダーは中央揃え、データは左上揃え
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_data = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # 全てのセルに対してフォントと配置を一律適用
        for row_idx_curr, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3), start=1):
            is_header = (row_idx_curr == 1)
            for cell in row:
                cell.font = font_header if is_header else font_data
                cell.alignment = align_header if is_header else align_data

        # 6. 列幅の最適化（JTF慣習および原本サンプルに準拠）
        ws.column_dimensions['A'].width = 26.13
        ws.column_dimensions['B'].width = 44.38
        ws.column_dimensions['C'].width = 110.0

        # 保存
        wb.save(filepath)
        print(f"Exported Excel to: {filepath}")
        return filepath
